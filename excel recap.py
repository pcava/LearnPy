# https://gaurav-adarshi.medium.com/working-with-excel-spreadsheets-in-python-2c3c2171879d

import pandas as pd
 
df = pd.DataFrame({'A':[1,2], 'B':[3,4]})
 
# creates the file from scratch removing all previous contents / sheets
df.to_excel("Input/ExcelTesting.xlsx", sheet_name='Sheet1', index=False)
 
# openpyxl
# creates the file from scratch removing all previous contents / sheets - mode: write
with pd.ExcelWriter("Input/ExcelTesting.xlsx", engine="openpyxl", mode="w") as writer:
    df.to_excel(writer, sheet_name='Sheet1', index=False)
   
# replace only the specific sheet removing previous content in it - mode: append
with pd.ExcelWriter("Input/ExcelTesting.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df.to_excel(writer, sheet_name='Sheet1', index=False)
 
# replace only the specific content in the specific sheet keeping previous content in it  - mode: append
with pd.ExcelWriter("Input/ExcelTesting.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    df.to_excel(writer, sheet_name='Sheet1', index=False)
   
# xlsxwriter
# creates the file from scratch removing all previous contents / sheets
with pd.ExcelWriter("Input/ExcelTesting.xlsx", engine="xlsxwriter") as writer:
    df.to_excel(writer, sheet_name='Sheet1', index=False)
 
# creates the file from scratch removing all previous contents / sheets - mode: write
with pd.ExcelWriter("Input/ExcelTesting.xlsx", engine="xlsxwriter", mode="w") as writer:
    df.to_excel(writer, sheet_name='Sheet1', index=False)
   
# Append mode is not supported with xlsxwriter!
# with pd.ExcelWriter("Input/ExcelTesting.xlsx", engine="xlsxwriter", mode="a", if_sheet_exists="replace") as writer:
#     df.to_excel(writer, sheet_name='Sheet1', index=False)
 
# Append mode is not supported with xlsxwriter!
# with pd.ExcelWriter("Input/ExcelTesting.xlsx", engine="xlsxwriter", mode="a", if_sheet_exists="overlay") as writer:
#     df.to_excel(writer, sheet_name='Sheet1', index=False)
 
###################################
 
# Delete previous data, add new data and refresh pivots
 
import pandas as pd
from pathlib import Path
 
file = "Input/ExcelTesting.xlsx"
filepath = Path.cwd() / file
# print(filepath)
 
df = pd.DataFrame({'a':[1,1], 'b':[1,1]})
df = pd.DataFrame({'a':[1,1,1,1], 'b':[1,1,1,1]})
 
# load excel file
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb = load_workbook(file)
ws = wb["Sheet1"]
 
# delete all rows after the 2nd one
# https://www.geeksforgeeks.org/how-to-delete-one-or-more-rows-in-excel-using-openpyxl/
while(ws.max_row > 2):
    ws.delete_rows(3)
 
# save
wb.save(file)
 
# add new data
with pd.ExcelWriter(file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    df.to_excel(writer, sheet_name='Sheet1', index=False)
 
#
def copy_formulas(file, src_sheet, src_cell, dst_range, dst_sheet = None):
   
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import rows_from_range
    from openpyxl.formula.translate import Translator
   
    wb = load_workbook(file)
 
    src_ws = wb[src_sheet]
    if dst_sheet is None:
        dst_ws = src_ws
    else:
        dst_ws = wb[dst_sheet]
   
    for row in rows_from_range(dst_range):
        for cell in row:
            dst_ws[cell] = Translator(src_ws[src_cell].value, src_cell).translate_formula(cell)
                   
    wb.save(file)
 
from openpyxl.utils import rows_from_range
for row in rows_from_range("A1:B1"):    
    for cell in row:
        copy_formulas(file, src_sheet = "Sheet2", src_cell = cell, dst_range = cell[0] + "2:" + cell[0] + "10")
 
wb.save(file)
# wb.close()
 
# refresh pivots
# https://docs.rackspace.com/blog/python-automate-everyday-tasks/
import win32com.client
xlapp = win32com.client.DispatchEx("Excel.Application")
xlapp.DisplayAlerts = False
# xlapp.Visible = True
wb = xlapp.Workbooks.Open(filepath)
wb.RefreshAll()
xlapp.CalculateUntilAsyncQueriesDone()
wb.Save()
wb.Close()
xlapp.Quit()
 
# copy-paste values
# must be done after phisically opening and saving the excel file
# https://stackoverflow.com/questions/23350581/openpyxl-1-8-5-reading-the-result-of-a-formula-typed-in-a-cell-using-openpyxl
wb = load_workbook(file, data_only=True)
ws = wb["Sheet2"]
for row in rows_from_range("A2:B10"):
    for cell in row:
        ws[cell].value = ws[cell].value
wb.save(file)
wb.close()
 
# test
test1 = pd.read_excel(file, "pivot1")
print(test1)
