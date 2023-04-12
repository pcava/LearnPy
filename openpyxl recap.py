# https://openpyxl.readthedocs.io/

# Sample Code
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("Output/sample.xlsx")


# https://openpyxl.readthedocs.io/en/stable/tutorial.html

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# create new worksheets
ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
# or
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
# or
ws3 = wb.create_sheet("Mysheet", -1) # insert at the penultimate position

# change the sheet name
ws.title = "New Title"

# get the sheet by using its name
ws3 = wb["New Title"]

# get the names of all sheets
print(wb.sheetnames)

# loop through sheets
for sheet in wb:
    print(sheet.title)

# create copies of worksheets within a single workbook
source = wb.active
target = wb.copy_worksheet(source)

# access a specific cell
c = ws['A4']

# assign a value to a specific cell
ws['A4'] = 4

# using the .cell() method
d = ws.cell(row=4, column=2, value=10)

# ranges of cells can be accessed using slicing
cell_range = ws['A1':'C2']

# ranges of rows or columns can be obtained similarly
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

# Worksheet.iter_rows() method
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
       print(cell)

# Likewise the Worksheet.iter_cols() method will return columns
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

# once we have a Cell, we can assign it a value
c.value = 'hello, world'
print(c.value)
d.value = 3.14
print(d.value)

# save a workbook - will overwrite existing files without warning!
wb.save('Output/balances.xlsx')

# load from a file
from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)

# Warning: openpyxl does currently not read all possible items in an 
# Excel file so shapes will be lost from existing files if they are 
# opened and saved with the same name.

# Close the workbook after reading
wb.close()


# Create a simple spreadsheet and bar chart
# https://openpyxl.readthedocs.io/en/stable/usage.html

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
treeData = [["Type", "Leaf Color", "Height"], 
            ["Maple", "Red", 549], 
            ["Oak", "Green", 783], 
            ["Pine", "Green", 1204]]

for row in treeData:
    ws.append(row)
    
from openpyxl.styles import Font
ft = Font(bold=True)
for row in ws["A1:C1"]:
    for cell in row:
        cell.font = ft

from openpyxl.chart import BarChart, Series, Reference
chart = BarChart()
chart.type = "col"
chart.title = "Tree Height"
chart.y_axis.title = 'Height (cm)'
chart.x_axis.title = 'Tree Type'
chart.legend = None

data = Reference(ws, min_col=3, min_row=2, max_row=4, max_col=3)
categories = Reference(ws, min_col=1, min_row=2, max_row=4, max_col=1)

chart.add_data(data)
chart.set_categories(categories)

ws.add_chart(chart, "E1")
wb.save("Output/TreeData.xlsx")


# Working with Pandas and NumPy
# https://openpyxl.readthedocs.io/en/stable/pandas.html
import pandas as pd
df = pd.DataFrame({'A':[1,2],'B':[3,4]})

# openpyxl.utils.dataframe.dataframe_to_rows() function 
# provides a simple way to work with Pandas Dataframe
from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
    
# to convert a worksheet to a Dataframe you can use the values property
import pandas as pd
df = pd.DataFrame(ws.values)



# Modify an existing excel file
# not best way
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )

df = pd.DataFrame({'Data': [10, 20, 30, 20, 15, 30, 45]})

wb = load_workbook("pandas_simple.xlsx")
del wb["Sheet1"]
ws = wb.create_sheet("Sheet1", 0)
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
wb.save("pandas_simple_updated.xlsx")

# Modify an existing excel file
# better way using mode='append'
from pathlib import Path
import shutil
source = Path('Output/pandas_simple.xlsx')
destination = Path('Output/pandas_simple_updated.xlsx')
shutil.copyfile(source, destination)

with pd.ExcelWriter("Output/pandas_simple_updated.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df.sort_values(by=['DataB'], ascending=False, inplace=True, ignore_index=True)
    df.to_excel(writer, sheet_name='Sheet1', index=False)

# Filtering and Sorting
filter_values = [20, 30]
wb = load_workbook("Output/pandas_simple_updated.xlsx")
ws = wb["Sheet1"]
ws.auto_filter.ref = ws.dimensions
col = FilterColumn(colId=0) # for column A
col.filters = Filters(filter=filter_values) # add selected values
ws.auto_filter.filterColumn.append(col) # add filter to the worksheet
for idx in (df.index[(~df['DataA'].isin(filter_values))].tolist()):
    ws.row_dimensions[idx+2].hidden = True
ws.auto_filter.add_sort_condition(f"B2:B{ws.max_row}", descending=True) # better to sort the original df directly
wb.save("Output/pandas_simple_updated.xlsx")


# Adding an empty row in the second line
import openpyxl
wb = openpyxl.load_workbook('Input/Test.xlsx')
ws = wb['Sheet1']
ws.insert_rows(2)
wb.save('Input/Test.xlsx')
